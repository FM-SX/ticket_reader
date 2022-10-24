from imbox import Imbox
from datetime import datetime, timedelta
import pandas as pd 
from barcode_reader import *
from openpyxl import Workbook
import os

username = open('login/username', 'r').read()
password = open('login/pass', 'r').read()
host = 'imap.gmail.com'
download_folder = "boletos"

mail = Imbox(host, username=username, password=password, ssl=True)

# Retorna todas as mensagens em um período de 30 dias que tenham anexos.
messages = mail.messages(date__gt=datetime.today() - timedelta(days=30), raw='has:attachment')

wb = Workbook()
ws = wb.active
r = 1
ws.cell(row=1, column=1).value = "Assunto"
ws.cell(row=1, column=2).value = "Código de barras"
ws.cell(row=1, column=3).value = "Linha digitável"
ws.cell(row=1, column=4).value = "Filename"

for (uid, message) in messages: 
    if len(message.attachments) > 0:
        for attach in message.attachments:
            att_file = attach["filename"]

            if '.pdf' in att_file:
                download_path = f"{download_folder}/{att_file}"

                with open(download_path, 'wb') as fp:
                    fp.write(attach['content'].read())

                try:
                    barcode = BarcodeReader(download_path)
                    linha_dig = linha_digitavel(barcode)
                except:
                    barcode = False

                if not barcode:
                    os.remove(download_path)
                
                else:
                    print(message.subject, '-', barcode)
                    r += 1
                    #Gravando no excel
                    ws.cell(row=r, column=1).value = message.subject
                    ws.cell(row=r, column=2).value = barcode
                    ws.cell(row=r, column=3).value = linha_dig
                    ws.cell(row=r, column=4).value = att_file

wb.save("boletos.xlsx")
mail.logout()